from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Nome do arquivo Excel
ARQUIVO_EXCEL = "dados.xlsx"

# Cria o arquivo Excel com cabeçalho se ele ainda não existir
if not os.path.exists(ARQUIVO_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nome", "Email", "Senha", "Igreja", "Cargo"])
    wb.save(ARQUIVO_EXCEL)

@app.route('/')
def index():
    return render_template('cadastro.html')

@app.route('/salvar', methods=['POST'])
def salvar():
    nome = request.form['nome']
    email = request.form['email']
    senha = request.form['senha']
    igreja = request.form['igreja']
    cargo = request.form.get('cargo', '')

    # Abre o arquivo Excel existente e adiciona uma nova linha
    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb.active
    ws.append([nome, email, senha, igreja, cargo])
    wb.save(ARQUIVO_EXCEL)

    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)